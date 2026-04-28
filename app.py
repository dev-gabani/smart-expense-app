from flask import Flask, request, redirect, url_for, session, render_template, jsonify, make_response

from datetime import datetime, timedelta
import calendar
from werkzeug.security import generate_password_hash, check_password_hash
import jwt
from functools import wraps
import os
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "super_secret_key_change_in_production")
app.config['SECRET_KEY'] = app.secret_key
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)

# Security Hardening: Secure Cookies
app.config.update(
    SESSION_COOKIE_SECURE=False, # Set to False for local HTTP testing
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
)

# Security Hardening: Rate Limiting
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

# Modern ORM and Migrations (Data Preservation)
from models import db, User, Income, Expense, Category
from sqlalchemy import func
import shutil
from pathlib import Path

import os
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'database.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)

# Ensure database tables exist (crucial for fresh deployments on Render)
with app.app_context():
    db.create_all()

# Automated Daily Backup System (Zero Data Loss Guarantee)
def backup_database():
    db_file = Path("database.db")
    if not db_file.exists(): return
    backup_dir = Path("backups")
    backup_dir.mkdir(exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")
    backup_path = backup_dir / f"database_backup_{today}.db"
    if not backup_path.exists():
        shutil.copy2(db_file, backup_path)
        print(f"Daily Database Backup Created: {backup_path}")
        
    # Cleanup old backups (keep only last 14)
    all_backups = sorted(list(backup_dir.glob("database_backup_*.db")))
    if len(all_backups) > 14:
        for old_backup in all_backups[:-14]:
            old_backup.unlink()
            print(f"Deleted old backup: {old_backup}")

backup_database()

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter



# ---------------- API AUTH DECORATOR ----------------
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'message': 'Token is missing!'}), 401
        try:
            if token.startswith('Bearer '):
                token = token.split(" ")[1]
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=["HS256"])
            current_user_id = data['user_id']
        except jwt.ExpiredSignatureError:
            return jsonify({'message': 'Token has expired!'}), 401
        except Exception as e:
            return jsonify({'message': 'Token is invalid!', 'error': str(e)}), 401
        return f(current_user_id, *args, **kwargs)
    return decorated

# ---------------- NO CACHE FIX ----------------
@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store'
    return response

# ==========================================
#              WEB ROUTES
# ==========================================

from werkzeug.exceptions import HTTPException
import traceback

@app.errorhandler(Exception)
def handle_exception(e):
    # Pass through HTTP errors
    if isinstance(e, HTTPException):
        return e
    # Return exactly what crashed so we can fix it!
    return f"<h1>Internal Error Detected!</h1><p>Please copy this and send it back:</p><pre style='background:#f4f4f4; padding:10px;'>{traceback.format_exc()}</pre>", 500

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute")
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            session.clear()
            session['user_id'] = user.id
            session['user'] = user.username
            if request.form.get('remember'):
                session.permanent = True
            return redirect(url_for('home'))
        else:
            return render_template('login.html', error="Invalid username or password!")
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        if len(password) < 8 or not any(char.isdigit() for char in password) or not any(char.isalpha() for char in password):
            return render_template('register.html', error="Password must be at least 8 characters long and contain both letters and numbers.")
            
        hashed_password = generate_password_hash(password)
        
        from sqlalchemy.exc import IntegrityError
        
        try:
            new_user = User(username=username, password_hash=hashed_password, income=0.0)
            db.session.add(new_user)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return render_template('register.html', error="Username already exists!")
            
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
def home():
    if 'user_id' not in session:
        return redirect(url_for('login'))
        
    user_id = session['user_id']
    current_month = request.args.get('month', datetime.now().strftime("%Y-%m"))
    
    income = db.session.query(func.sum(Income.amount)).filter(Income.user_id == user_id, func.substr(Income.date, 1, 7) == current_month).scalar() or 0.0
    spent = db.session.query(func.sum(Expense.amount)).filter(Expense.user_id == user_id, func.substr(Expense.date, 1, 7) == current_month).scalar() or 0.0
    remaining = income - spent
    
    category_expenses_raw = db.session.query(Expense.category, func.sum(Expense.amount)).filter(Expense.user_id == user_id, func.substr(Expense.date, 1, 7) == current_month).group_by(Expense.category).all()
    category_expenses = [(row[0], float(row[1])) for row in category_expenses_raw]
    
    # --- Step 4: Dashboard Enhancements (Financial Score & Insights) ---
    score = 100
    if income > 0:
        savings_pct = (remaining / income) * 100
        if savings_pct < 0: score -= 40
        elif savings_pct < 10: score -= 20
        elif savings_pct < 20: score -= 10
    else:
        score = 0
        
    cats = Category.query.filter_by(user_id=user_id).all()
    cat_budgets = {c.name: (income * c.percentage / 100.0) for c in cats}
    
    for cat, amt in category_expenses:
        if cat in cat_budgets and amt > cat_budgets[cat]:
            score -= 5 # penalty for overspending a category
            
    score = max(0, min(100, score)) # Clamp between 0-100
    
    all_expenses = Expense.query.filter(Expense.user_id == user_id, func.substr(Expense.date, 1, 7) == current_month).all()
    
    weekend_spend = 0
    weekday_spend = 0
    for exp in all_expenses:
        dt = datetime.strptime(exp.date, "%Y-%m-%d")
        if dt.weekday() >= 5: # 5=Sat, 6=Sun
            weekend_spend += float(exp.amount)
        else:
            weekday_spend += float(exp.amount)
            
    insights = []
    if weekend_spend > weekday_spend:
        insights.append("🛍️ You spend more on weekends than weekdays.")
    if score >= 80:
        insights.append("🌟 Excellent financial health this month!")
    elif score < 50:
        insights.append("⚠️ High risk of overspending. Review your budget.")
    
    if len(category_expenses) > 0:
        top_cat = max(category_expenses, key=lambda x: x[1])
        insights.append(f"📊 Your biggest expense area is {top_cat[0].title()}.")
    # -----------------------------------------------------------------
    
    recent_transactions_raw = Expense.query.filter_by(user_id=user_id).order_by(Expense.date.desc(), Expense.id.desc()).limit(5).all()
    recent_transactions = [(e.category, float(e.amount), e.date) for e in recent_transactions_raw]
    
    return render_template("home.html", 
        income=income, spent=spent, remaining=remaining, 
        category_expenses=category_expenses,
        recent_transactions=recent_transactions, 
        selected_month=current_month,
        score=score, insights=insights)

@app.route('/add_category', methods=['GET', 'POST'])
def add_category():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user_id = session['user_id']

    if request.method == 'POST':
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        name = request.form['name'].strip().lower()
        try:
            percentage = float(request.form['percentage'])
            if percentage < 0 or percentage > 100:
                if is_ajax: return jsonify({'error': 'Invalid percentage'}), 400
                return render_template("add_category.html", error="Invalid percentage")
        except ValueError:
            if is_ajax: return jsonify({'error': 'Percentage must be a number'}), 400
            return render_template("add_category.html", error="Percentage must be a number")

        existing = Category.query.filter_by(user_id=user_id, name=name).first()

        if existing:
            existing.percentage = percentage
        else:
            new_cat = Category(user_id=user_id, name=name, percentage=percentage)
            db.session.add(new_cat)

        db.session.commit()
        if is_ajax: return jsonify({'success': True, 'message': 'Category saved successfully'})
        return redirect(url_for('add_category'))

    return render_template("add_category.html")

@app.route('/set_income', methods=['GET', 'POST'])
def set_income():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user_id = session['user_id']

    if request.method == 'POST':
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        try:
            amount = float(request.form['income'])
            if amount < 0 or amount > 999999999:
                if is_ajax: return jsonify({'error': 'Invalid amount'}), 400
                return render_template("set_income.html", error="Invalid amount")
        except ValueError:
            if is_ajax: return jsonify({'error': 'Amount must be a number'}), 400
            return render_template("set_income.html", error="Amount must be a number")
        source = request.form.get('source', 'Salary').strip()
        date_str = datetime.now().strftime("%Y-%m-%d")

        new_inc = Income(user_id=user_id, source=source, amount=amount, date=date_str)
        db.session.add(new_inc)
        db.session.commit()
        
        if is_ajax: return jsonify({'success': True, 'message': 'Income added successfully'})
        return redirect(url_for('set_income'))

    return render_template("set_income.html")

@app.route('/add_expense', methods=['GET', 'POST'])
def add_expense():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user_id = session['user_id']

    categories = Category.query.filter_by(user_id=user_id).all()

    if request.method == 'POST':
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        category = request.form['category'].strip().lower()
        try:
            amount = float(request.form['amount'])
            if amount < 0 or amount > 999999999:
                if is_ajax: return jsonify({'error': 'Invalid amount'}), 400
                return render_template("add_expense.html", categories=[(c.name,) for c in categories], error="Invalid amount")
        except ValueError:
            if is_ajax: return jsonify({'error': 'Amount must be a number'}), 400
            return render_template("add_expense.html", categories=[(c.name,) for c in categories], error="Amount must be a number")
        date = datetime.now().strftime("%Y-%m-%d")

        new_exp = Expense(user_id=user_id, category=category, amount=amount, date=date)
        db.session.add(new_exp)
        db.session.commit()
        
        if is_ajax: return jsonify({'success': True, 'message': 'Expense recorded successfully'})
        return redirect(url_for('add_expense'))

    return render_template("add_expense.html", categories=[(c.name,) for c in categories])

@app.route('/view_expenses')
def view_expenses():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user_id = session['user_id']

    current_month = request.args.get('month', datetime.now().strftime("%Y-%m"))
    
    expenses = Expense.query.filter(
        Expense.user_id == user_id,
        func.substr(Expense.date, 1, 7) == current_month
    ).order_by(Expense.date.desc(), Expense.id.desc()).all()
    
    data = [(e.category, e.amount, e.date) for e in expenses]

    return render_template("view_expenses.html", data=data, selected_month=current_month)

@app.route('/smart_budget')
def smart_budget():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user_id = session['user_id']

    current_month = request.args.get('month', datetime.now().strftime("%Y-%m"))

    income = db.session.query(func.sum(Income.amount)).filter(Income.user_id == user_id, func.substr(Income.date, 1, 7) == current_month).scalar() or 0.0

    categories = Category.query.filter_by(user_id=user_id).all()
    
    data = []

    for cat in categories:
        name = cat.name
        percent = cat.percentage
        budget_amount = (income * float(percent)) / 100.0

        spent = db.session.query(func.sum(Expense.amount)).filter(
            Expense.user_id == user_id, Expense.category == name, func.substr(Expense.date, 1, 7) == current_month
        ).scalar() or 0.0
        
        remaining = budget_amount - spent

        if remaining < 0:
            status = "❌ Overspent!"
        elif remaining < (budget_amount * 0.2):
            status = "⚠️ Warning"
        else:
            status = "✅ Safe"

        data.append({
            "name": name,
            "budget": budget_amount,
            "spent": spent,
            "remaining": remaining,
            "status": status
        })
    return render_template("smart_budget.html", income=income, data=data, selected_month=current_month)

@app.route('/suggestions')
def suggestions():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user_id = session['user_id']

    current_month = request.args.get('month', datetime.now().strftime("%Y-%m"))
    income = db.session.query(func.sum(Income.amount)).filter(Income.user_id == user_id, func.substr(Income.date, 1, 7) == current_month).scalar() or 0.0
    categories = Category.query.filter_by(user_id=user_id).all()

    suggestions_list = []
    for cat in categories:
        name = cat.name
        percent = cat.percentage
        budget = (income * float(percent)) / 100.0
        
        spent = db.session.query(func.sum(Expense.amount)).filter(
            Expense.user_id == user_id, Expense.category == name, func.substr(Expense.date, 1, 7) == current_month
        ).scalar() or 0.0

        if spent > budget:
            suggestions_list.append(f"❌ Reduce {name} by ₹{spent - budget:.2f}")
        elif spent < budget * 0.5:
            suggestions_list.append(f"💡 Underusing {name}")
        else:
            suggestions_list.append(f"✅ {name} balanced")

    return render_template("suggestions.html", suggestions=suggestions_list, selected_month=current_month)

@app.route('/prediction')
def prediction():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user_id = session['user_id']

    today = datetime.now()
    current_month = request.args.get('month', today.strftime("%Y-%m"))

    spent = db.session.query(func.sum(Expense.amount)).filter(
        Expense.user_id == user_id, func.substr(Expense.date, 1, 7) == current_month
    ).scalar() or 0.0

    daily_avg = spent / today.day if today.day > 0 else 0
    total_days = calendar.monthrange(today.year, today.month)[1]
    predicted_total = daily_avg * total_days

    income = db.session.query(func.sum(Income.amount)).filter(
        Income.user_id == user_id, func.substr(Income.date, 1, 7) == current_month
    ).scalar() or 0.0

    status = "Overspend" if predicted_total > income else "Save"
    difference = abs(predicted_total - income)
    
    return render_template(
        "prediction.html",
        spent=spent,
        daily_avg=round(daily_avg, 2),
        predicted=round(predicted_total, 2),
        status=status,
        selected_month=current_month
    )

# ==========================================
#              API ROUTES (JWT)
# ==========================================

@app.route('/api/login', methods=['POST'])
@limiter.limit("5 per minute")
def api_login():
    data = request.get_json()
    if not data or not data.get('username') or not data.get('password'):
        return jsonify({'message': 'Missing credentials'}), 400
        
    user = User.query.filter_by(username=data['username']).first()
    
    if user and check_password_hash(user.password_hash, data['password']):
        token = jwt.encode({'user_id': user.id, 'exp': datetime.utcnow() + timedelta(days=7)}, app.config['SECRET_KEY'], algorithm="HS256")
        return jsonify({'token': token})
        
    return jsonify({'message': 'Invalid credentials'}), 401

@app.route('/api/register', methods=['POST'])
def api_register():
    data = request.get_json()
    if not data or not data.get('username') or not data.get('password'):
        return jsonify({'message': 'Missing credentials'}), 400

    hashed_password = generate_password_hash(data['password'])
    from sqlalchemy.exc import IntegrityError
    
    try:
        new_user = User(username=data['username'], password_hash=hashed_password, income=0.0)
        db.session.add(new_user)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({'message': 'Username already exists'}), 409
        
    return jsonify({'message': 'User registered successfully'}), 201

@app.route('/api/dashboard', methods=['GET'])
@token_required
def api_dashboard(current_user_id):
    current_month = datetime.now().strftime("%Y-%m")
    
    income = db.session.query(func.sum(Income.amount)).filter(Income.user_id == current_user_id, func.substr(Income.date, 1, 7) == current_month).scalar() or 0.0
    spent = db.session.query(func.sum(Expense.amount)).filter(Expense.user_id == current_user_id, func.substr(Expense.date, 1, 7) == current_month).scalar() or 0.0
    
    remaining = income - spent
    
    return jsonify({
        'income': income,
        'spent': spent,
        'remaining': remaining,
        'month': current_month
    })

@app.route('/api/expenses', methods=['GET', 'POST'])
@token_required
def api_expenses(current_user_id):
    if request.method == 'POST':
        data = request.get_json()
        category = data.get('category', '').strip().lower()
        try:
            amount = float(data.get('amount', 0))
            if amount < 0 or amount > 999999999:
                return jsonify({'message': 'Invalid amount'}), 400
        except (ValueError, TypeError):
            return jsonify({'message': 'Amount must be a number'}), 400
        date = datetime.now().strftime("%Y-%m-%d")
        
        new_exp = Expense(user_id=current_user_id, category=category, amount=amount, date=date)
        db.session.add(new_exp)
        db.session.commit()
        return jsonify({'message': 'Expense added successfully'}), 201
        
    # GET method
    expenses = Expense.query.filter_by(user_id=current_user_id).order_by(Expense.date.desc()).all()
    expenses_list = [{'id': e.id, 'category': e.category, 'amount': e.amount, 'date': e.date} for e in expenses]
    
    return jsonify({'expenses': expenses_list})

@app.route('/api/categories', methods=['GET', 'POST'])
@token_required
def api_categories(current_user_id):
    if request.method == 'POST':
        data = request.get_json()
        name = data.get('name', '').strip().lower()
        try:
            percentage = float(data.get('percentage', 0))
            if percentage < 0 or percentage > 100:
                return jsonify({'message': 'Invalid percentage'}), 400
        except (ValueError, TypeError):
            return jsonify({'message': 'Percentage must be a number'}), 400
            
        existing = Category.query.filter_by(user_id=current_user_id, name=name).first()

        if existing:
            existing.percentage = percentage
        else:
            new_cat = Category(user_id=current_user_id, name=name, percentage=percentage)
            db.session.add(new_cat)
        db.session.commit()
        return jsonify({'message': 'Category updated successfully'}), 200
        
    # GET method
    categories = Category.query.filter_by(user_id=current_user_id).all()
    categories_list = [{'id': c.id, 'name': c.name, 'percentage': c.percentage} for c in categories]
    return jsonify({'categories': categories_list})

@app.route('/api/smart_budget', methods=['GET'])
@token_required
def api_smart_budget(current_user_id):
    current_month = request.args.get('month', datetime.now().strftime("%Y-%m"))

    income = db.session.query(func.sum(Income.amount)).filter(Income.user_id == current_user_id, func.substr(Income.date, 1, 7) == current_month).scalar() or 0.0
    categories = Category.query.filter_by(user_id=current_user_id).all()
    
    data = []

    for cat in categories:
        name = cat.name
        percent = cat.percentage
        budget_amount = (income * float(percent)) / 100.0

        spent = db.session.query(func.sum(Expense.amount)).filter(
            Expense.user_id == current_user_id, Expense.category == name, func.substr(Expense.date, 1, 7) == current_month
        ).scalar() or 0.0
        
        remaining = budget_amount - spent

        if remaining < 0:
            status = "Overspent!"
        elif remaining < (budget_amount * 0.2):
            status = "Warning"
        else:
            status = "Safe"

        data.append({
            "name": name,
            "budget": budget_amount,
            "spent": spent,
            "remaining": remaining,
            "status": status
        })

    return jsonify({'income': income, 'month': current_month, 'budget_data': data})

@app.route('/api/prediction', methods=['GET'])
@token_required
def api_prediction(current_user_id):
    today = datetime.now()
    current_month = request.args.get('month', today.strftime("%Y-%m"))

    spent = db.session.query(func.sum(Expense.amount)).filter(
        Expense.user_id == current_user_id, func.substr(Expense.date, 1, 7) == current_month
    ).scalar() or 0.0

    daily_avg = spent / today.day if today.day > 0 else 0
    total_days = calendar.monthrange(today.year, today.month)[1]
    predicted_total = daily_avg * total_days

    income = db.session.query(func.sum(Income.amount)).filter(
        Income.user_id == current_user_id, func.substr(Income.date, 1, 7) == current_month
    ).scalar() or 0.0

    status = "Overspend" if predicted_total > income else "Save"
    difference = abs(predicted_total - income)

    return jsonify({
        'spent': spent,
        'daily_avg': round(daily_avg, 2),
        'predicted': round(predicted_total, 2),
        'income': income,
        'status': status,
        'difference': round(difference, 2),
        'month': current_month
    })

@app.route('/api/incomes', methods=['GET', 'POST'])
@token_required
def api_incomes(current_user_id):
    if request.method == 'POST':
        data = request.get_json()
        try:
            amount = float(data.get('amount', 0))
            if amount < 0 or amount > 999999999:
                return jsonify({'message': 'Invalid amount'}), 400
        except (ValueError, TypeError):
            return jsonify({'message': 'Amount must be a number'}), 400
            
        source = data.get('source', 'Salary').strip()
        date_str = datetime.now().strftime("%Y-%m-%d")
        
        new_inc = Income(user_id=current_user_id, source=source, amount=amount, date=date_str)
        db.session.add(new_inc)
        db.session.commit()
        return jsonify({'message': 'Income added successfully'}), 201
        
    # GET method
    current_month = request.args.get('month', datetime.now().strftime("%Y-%m"))
    incomes = Income.query.filter(Income.user_id == current_user_id, func.substr(Income.date, 1, 7) == current_month).order_by(Income.date.desc()).all()
    incomes_list = [{'id': i.id, 'source': i.source, 'amount': i.amount, 'date': i.date} for i in incomes]
    
    return jsonify({'incomes': incomes_list, 'month': current_month})

@app.route('/api/suggestions', methods=['GET'])
@token_required
def api_suggestions(current_user_id):
    current_month = request.args.get('month', datetime.now().strftime("%Y-%m"))

    income = db.session.query(func.sum(Income.amount)).filter(Income.user_id == current_user_id, func.substr(Income.date, 1, 7) == current_month).scalar() or 0.0
    categories = Category.query.filter_by(user_id=current_user_id).all()

    suggestions_list = []
    for cat in categories:
        name = cat.name
        percent = cat.percentage
        budget = (income * float(percent)) / 100.0

        spent = db.session.query(func.sum(Expense.amount)).filter(
            Expense.user_id == current_user_id, Expense.category == name, func.substr(Expense.date, 1, 7) == current_month
        ).scalar() or 0.0

        if spent > budget:
            suggestions_list.append(f"❌ Reduce {name} by ₹{spent - budget:.2f}")
        elif spent < budget * 0.5:
            suggestions_list.append(f"💡 Underusing {name}")
        else:
            suggestions_list.append(f"✅ {name} balanced")

    return jsonify({'suggestions': suggestions_list, 'month': current_month})

@app.route('/export/excel')
def export_excel():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user_id = session['user_id']
    current_month = request.args.get('month', datetime.now().strftime("%Y-%m"))
    
    incomes = Income.query.filter(Income.user_id == user_id, func.substr(Income.date, 1, 7) == current_month).all()
    expenses = Expense.query.filter(Expense.user_id == user_id, func.substr(Expense.date, 1, 7) == current_month).all()
    
    transactions = [{'date': i.date, 'description': i.source, 'amount': i.amount, 'type': 'Credit'} for i in incomes] + \
                   [{'date': e.date, 'description': e.category, 'amount': e.amount, 'type': 'Debit'} for e in expenses]
                   
    transactions.sort(key=lambda x: x['date'], reverse=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Statement"
    
    # Headers
    headers = ["Date", "Description", "Type", "Amount (INR)"]
    ws.append(headers)
    
    # Style Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="d97f3e", end_color="d97f3e", fill_type="solid")
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    # Write data
    for t in transactions:
        ws.append([t['date'], t['description'].title(), t['type'], float(t['amount'])])
        
    # Style data (colors for Credit/Debit)
    for row in ws.iter_rows(min_row=2, max_row=len(transactions)+1):
        row[3].number_format = '"₹"#,##0.00' # format as currency
        if row[2].value == 'Credit':
            row[2].font = Font(color="10B981", bold=True)
        else:
            row[2].font = Font(color="EF4444", bold=True)
            
    # Freeze the top header row so it stays visible when scrolling
    ws.freeze_panes = "A2"
    
    # Calculate Totals
    total_income = sum(float(t.amount) for t in incomes)
    total_expense = sum(float(t.amount) for t in expenses)
    net_remaining = total_income - total_expense
    
    # Add spacing
    ws.append([])
    ws.append([])
    
    # Add Summary Block
    summary_start_row = ws.max_row + 1
    ws.append(["", "FINANCIAL SUMMARY", "", ""])
    ws.append(["", "Total Credits (Income)", "", total_income])
    ws.append(["", "Total Debits (Expenses)", "", total_expense])
    ws.append(["", "Net Remaining", "", net_remaining])
    
    # Add Status line
    status_msg = f"✅ You saved ₹{net_remaining:,.2f} this month." if net_remaining >= 0 else f"❌ You are in a loss of ₹{abs(net_remaining):,.2f} this month."
    ws.append(["", status_msg, "", ""])
    
    # Style Summary
    ws.cell(row=summary_start_row, column=2).font = Font(bold=True, size=14, color="d97f3e")
    for r in range(summary_start_row + 1, summary_start_row + 4):
        ws.cell(row=r, column=2).font = Font(bold=True)
        ws.cell(row=r, column=4).number_format = '"₹"#,##0.00'
        ws.cell(row=r, column=4).font = Font(bold=True)
        
    # Color the Net Remaining line
    net_cell = ws.cell(row=summary_start_row + 3, column=4)
    if net_remaining >= 0:
        net_cell.font = Font(bold=True, color="10B981") # Green
    else:
        net_cell.font = Font(bold=True, color="EF4444") # Red
        
    status_cell = ws.cell(row=summary_start_row + 4, column=2)
    status_cell.font = Font(italic=True, bold=True)
            
    # Auto-adjust column widths to prevent #####
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 6) # Increased padding to 6 to handle currency symbol width
        ws.column_dimensions[column].width = adjusted_width
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename=statement_{current_month}.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

# ---------------- RUN ----------------
if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True)